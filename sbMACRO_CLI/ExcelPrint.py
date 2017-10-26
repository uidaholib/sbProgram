import g
import os
from pandas import DataFrame
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

from pprint import pprint

import requests
import json
import pysb


sb = pysb.SbSession()

def main():
    #try:
    PossiblePermissionsIssuesURL = []
    PossiblePermissionsIssuesURL[:] = []
    MissingDataURL= []
    MissingDataURL[:] = []
    if g.Exceptions != []:
        for i in g.Exceptions:
            json = sb.get_item(i)
            PossiblePermissionsIssuesURL.append(json['link']['url'])
    if g.MissingData != []:
        for i in g.MissingData:
            json = sb.get_item(i)
            MissingDataURL.append(json['link']['url'])

    wb = Workbook()
    ws = wb.active
    ws.title = "Report"

    if g.MissingData != []:
        ws_missing = wb.create_sheet("Missing", 1) #Inserts a new sheet, named "Missing" in the second position
        #ws_missing.sheet_properties.tabColor = "1072BA" #Makes the tab for the sheet red so it draws attention.
    if g.Exceptions != []:
        ws_exceptions = wb.create_sheet("Exceptions", 2) #Inserts a new sheet, named "Exceptions" in the third position
        #ws_exceptions.sheet_properties.tabColor = "1072BA" #Makes the tab for the sheet red so it draws attention.


    df = DataFrame({'ID': g.ID, 'Object Type': g.ObjectType, 'Name': g.Name,
                    'Fiscal Year': g.FiscalYear, 'Project': g.Project,
                    'Data in Project (GB)': g.DataInProject,
                    'Data per File (KB)': g.DataPerFile,
                    'Fiscal Year Total Data (GB)': g.totalFYData,
                    'Running Data Total (GB)': g.RunningDataTotal,
                        })
                        #include these eventually: 'Missing Data?': L7MissingData, 'Exceptions/Permissions Issues': L9Exceptions

    dfOrdered = df[['ID', 'Object Type', 'Name', 'Fiscal Year', 'Project',
                'Data in Project (GB)', 'Data per File (KB)',
                'Fiscal Year Total Data (GB)', 'Running Data Total (GB)',
                ]]
                #include these eventually: 'Missing Data?', 'Exceptions/Permissions Issues'
    #if g.MissingData != []:
    df_missing = DataFrame({'Missing Data ID': g.MissingData,
                            'Missing Data URL': MissingDataURL})
    df_missing = df_missing[['Missing Data ID', 'Missing Data URL']]

    #if g.Exceptions != []:
    df_exceptions = DataFrame({'Exception/Permission Issue ID': g.Exceptions,
                               'Exception/Permission Issue URL':
                               PossiblePermissionsIssuesURL}) #include 'Exception ID': L10Exceptions_IDs, later
    df_exceptions = df_exceptions[['Exception/Permission Issue ID',
                                              'Exception/Permission Issue URL']] #include 'Exception ID' later


    for r in dataframe_to_rows(dfOrdered, index=False, header=True):
        ws.append(r)

    for cell in ws[1]:
        cell.style = 'Pandas'

    if MissingDataURL != []:
        for r in dataframe_to_rows(df_missing, index=False, header=True):
            ws_missing.append(r)

        for cell in ws_missing[1]:
            cell.style = 'Pandas'

    if g.Exceptions != []:
        for r in dataframe_to_rows(df_exceptions, index=False, header=True):
            ws_exceptions.append(r)

        for cell in ws_exceptions[1]:
            cell.style = 'Pandas'

    print(df)
    print('''

    ---------Let\'s try reordering that...

    ''')
    print(dfOrdered)
    if g.Exceptions != [] or g.MissingData != []:
        print('''

    Items missing data:
              ''')
        print(df_missing)
        print('''

    Exceptions raised:
              ''')
        print(df_exceptions)
    print('''
    Does that look correct?
    (Y / N)
    ''')
    correct = input("> ").lower()
    if 'y' in correct:
        ask(wb)
    elif 'n' in correct:
        import editGPY
        editGPY.main()
        main()

def ask(wb):
    print ('''

    Would you like to save this?
    (Y / N)''')
    answer = input('> ').lower()
    if 'y' in answer:
        print('''
    Where would you like to save the file? Copy and paste a file path or '''+
    '''type "Desktop" to save it to the desktop (If you run Linux, you must'''+
    ''' paste a path).''')
        answer2 = input("> ").lower()
        if 'desk' in answer2:
            filePath = os.path.expanduser("~/Desktop/")
            saveExcel(wb, filePath)
        elif '/' in answer2 or '\\' in answer2:
            filePath = answer2
            saveExcel(wb, filePath)
        else:
            print('''
    That didn't appear to be a path. Please try again.''')
            ask(wb)

    elif 'n' in answer:
        print('''
        Ok, we won't save it.''')
        return
#    except (ValueError, Exception) as e:
#        print('''
#    ----------------------------WARNING: Something went wrong in the function "main" in ExcelPrint.py.''')
#        exit()


def saveExcel(wb, filePath):
    ChosenFiscalYear = g.FiscalYear[-1]  # Most recent Fiscal Year.
    print('''
    Would you like to name it something other than "'''+str(ChosenFiscalYear)+
    ''' Data Metrics.xlsx"?
    Beware: chosing the same name as an existing file overwrites that file '''+
    '''without warning.''')
    answer = input('> ').lower()
    if 'y' in answer or 'other' in answer:
        print('''
    What name would you like to give the Excel file?''')
        name = input('> ')
        wb.save(str(filePath)+str(name)+".xlsx")
        print('''
    Workbook saved as "'''+str(name)+'''.xlsx" in \''''+str(filePath)+'''\'.''')
        return
    elif 'n' in answer or 'keep' in answer:
        wb.save(str(filePath)+str(ChosenFiscalYear)+" Data Metrics.xlsx")
        print('''
    Workbook saved as "'''+str(ChosenFiscalYear)+''' Data Metrics.xlsx" in \''''+str(filePath)+'''\'.''')
        return
    else:
        print('''
    I'm sorry, I didn't get that. Please type 'y', 'other', 'n', or 'keep'
    ''')
        saveExcel(wb, filePath)
