"""Define main application routes."""
import os
import sys
import openpyxl
from datetime import datetime
from collections import OrderedDict
from flask import render_template, redirect, url_for, request, jsonify, current_app, session
from flask_login import current_user, login_required
from wtforms import StringField, SubmitField, TextAreaField, PasswordField
from wtforms import BooleanField
from wtforms.validators import ValidationError, DataRequired, Length, Email
from wtforms.validators import Optional
from app import db
from lxml import etree
import urllib
import requests
from xml.dom import minidom


from app.main.forms import EditProfileForm, FyForm, SearchForm
from app.models import User, casc, FiscalYear, Project, Item, SbFile, MasterDetails

from app.main import bp
from app.main.metadata import write_metadata
from app.main.forms import EditProfileForm, FyForm, GeneralForm
from app.models import User, casc, FiscalYear, Project, Item, SbFile
from app.auth.read_sheets import get_sheet_name, parse_values
from app.updater.__init__ import casc_update, search_table_update, graphs_update, proj_matches_update
import multiprocessing
import time

from nltk.corpus import stopwords

from config import Config
# from sbmacro import socketio
from pprint import pprint


# my_root = os.path.dirname(os.path.abspath(__file__))
# path_to_static = os.path.join(my_root, 'templates/static/')
path_to_static = os.getcwd() + '/app/main/templates/static/'


@bp.before_app_request
def before_request():
    """Update user 'last seen' field before each request."""
    if current_user.is_authenticated:
        current_user.last_seen = datetime.utcnow()
        db.session.commit()
        current_user.search_form = SearchForm()

@bp.route('/', methods = ['GET', 'POST'])  # Also accepts
@bp.route('/index', methods = ['GET', 'POST'])  # Default
def index():
    # class F(GeneralForm):
    #     def __init__(i, buttonText):
    #         super(F, i).__init__()
    #         # i.name = BooleanField('static field')
    #         i.submit = SubmitField(buttonText)

    # form = F('Refresh Metadata')

    class F(FyForm):
        pass

    form = F()

    """Render splash page for sbMACRO."""
    return(render_template('index.html', **locals(), title = "Welcome to sbMACRO"))


@bp.route('/metadata', methods = ['GET', 'POST'])
def metadata():

    tag_to_search = request.form['tag_to_search']
    custom_stopwords = ['climate', 'change']
    protocol = 'xml'
    url_file_name = 'metadata_urls.csv'
    us_states_file_name = 'us-states.csv'

    us_states = []
    with open(path_to_static + us_states_file_name, 'r') as file:
        for state in file:
            us_states.append(state.strip().lower())

    stop_words = us_states + stopwords.words('english') + custom_stopwords

    if request.method == 'POST':
        casc_name = request.form['casc_name']

        metadata_urls = []

        with open(path_to_static + url_file_name, 'r') as file:
            for line in file:
                casc, url = line.split(',')
                if casc == casc_name:
                    metadata_urls.append(url)

        # write to csv to be read by wordcloud module
        write_metadata(casc_name, tag_to_search, metadata_urls, stop_words)

    return ''


@bp.route('/fiscal_years')
@bp.route('/fiscalyears')
@bp.route('/select_fiscalyear', methods=['GET', 'POST'])  # Also accepts
@bp.route('/select_fiscalyears', methods=['GET', 'POST'])  # Default
def fiscalyear():
    """Retrieve Fiscal Years and display for selection by user."""
    cascs = db.session.query(casc).order_by(casc.name).all()
    cascs_and_fys = {}

    class F(FyForm):
        pass

    list_fy = []
    for curr_casc in cascs:
        cascs_and_fys[curr_casc.name] = {}
        cascs_and_fys[curr_casc.name]["id"] = curr_casc.id
        fys = db.session.query(FiscalYear).order_by(
            FiscalYear.name).filter(
            FiscalYear.casc_id == curr_casc.id).all()
        cascs_and_fys[curr_casc.name]["fiscal_years"] = []
        for fy in fys:
            fiscal_year = {}
            list_fy.append("fy" + str(fy.id))
            fiscal_year["id"] = fy.id
            fiscal_year["name"] = fy.name
            cascs_and_fys[curr_casc.name]["fiscal_years"].append(fiscal_year)
            setattr(F, "fy" + str(fy.id), BooleanField(fy.name))

    form = F()
    if form.validate_on_submit():
        id_list = []
        projects = []
        for fy in list_fy:
            fy_attr = getattr(form, fy)
            selected = fy_attr.data
            if selected:
                id_list.append(fy.replace("fy", ""))

        for i in id_list:
            fy_model = db.session.query(FiscalYear).get(i)
            for proj in fy_model.projects:
                project_dict = {}
                project_dict['fy_id'] = i
                project_dict['casc_id'] = fy_model.casc_id
                project_dict['proj_id'] = proj.id
                projects.append(project_dict)

        session["projects"] = projects
        return redirect(url_for('main.report'))
    elif request.method == 'GET':
        pass

    return render_template('fiscalYears.html',
                           form=form,
                           cascs_and_fys=cascs_and_fys,
                           title="Select Fiscal Years"),400


@bp.route('/update_db', methods=['GET', 'POST'])
def update_db():
    """Retrieve CASCs and display for selection by user."""

    update_graphs = False
    update_search_table = False
    update_proj_matches = False
    cascs_to_update = []

    list_of_cascs = ['Alaska', 'North Central', 'Northeast', 'Northwest', 'Pacific', 'South Central', 'Southeast', 'Southwest', 'National']
    
    class F(FyForm):
        pass

    # set form attributs for 'update_search_table' checkbox
    setattr(F, str('update_search_table'), BooleanField('update_search_table'))
    # set form attributs for 'update_graphs' checkbox
    setattr(F, str('update_graphs'), BooleanField('update_graphs'))
    # set form attributs for 'update_proj_matches' checkbox
    setattr(F, str('update_proj_matches'), BooleanField('update_proj_matches'))
    # set form attributs for casc checkboxes
    for curr_casc in list_of_cascs:
        setattr(F, str(curr_casc), BooleanField(curr_casc))

    form = F()
    if form.validate_on_submit():

        if getattr(form, 'update_search_table').data:
            update_search_table = True
        if getattr(form, 'update_graphs').data:
            update_graphs = True
        if getattr(form, 'update_proj_matches').data:
            update_proj_matches = True
        
        for csc in list_of_cascs:
            csc_attr = getattr(form, csc)

            selected = csc_attr.data
            if selected:
                cascs_to_update.append(csc)

        session['update_graphs'] = update_graphs
        session['update_search_table'] = update_search_table
        session['update_proj_matches'] = update_proj_matches
        session['cascs_to_update'] = cascs_to_update

        return redirect(url_for('main.updates'))

    elif request.method == 'GET':
        pass

    return render_template('update_db.html', form = form, list_of_cascs = list_of_cascs),400

# @socketio.on('connect', namespace='/test')
@bp.route('/updates')
def updates():
    """Refresh master details table and update the cascs selected for update"""

    # --- select where to update from ---
    source = 'sciencebase'
    # source = 'file'

    update_graphs = session['update_graphs']
    if update_graphs:
        print('Starting graph update thread')
        graph_upate_thread = multiprocessing.Process(target = graphs_update)
        graph_upate_thread.start()

    update_search_table = session['update_search_table']
    if update_search_table:
        print('Starting search table update thread')
        search_table_update_thread = multiprocessing.Process(target = search_table_update, args = (source,))
        search_table_update_thread.start()

    update_proj_matches = session['update_proj_matches']
    if update_proj_matches:
        print('Starting project matches update thread')
        proj_matches_update_thread = multiprocessing.Process(target = proj_matches_update)
        proj_matches_update_thread.start()

    cascs_to_update = session['cascs_to_update']
    if cascs_to_update:
        print('Starting CASC updates...')
        casc_update_thread = multiprocessing.Process(target = casc_update, args = (cascs_to_update,))
        casc_update_thread.start()

    return render_template("updates.html", update_graphs = update_graphs, update_search_table = update_search_table, update_proj_matches = update_proj_matches, cascs_to_update = cascs_to_update)


@bp.route('/casc_projects/<params>', methods = ['GET', 'POST'])
def casc_projects(params):

    casc_name, num_projects, num_datasets = params.split('|')

    return render_template('casc_projects.html', casc_name = casc_name, num_projects = num_projects, num_datasets = num_datasets)


@bp.route('/proj_compare', methods = ['GET', 'POST'])
def proj_compare():

    return render_template('proj_compare.html')


@bp.route('/projects', methods = ['GET', 'POST'])
@bp.route('/select_project', methods = ['GET', 'POST'])
@bp.route('/select_projects', methods = ['GET', 'POST'])
def project():
    """Display and implement selection/searching for projects by URL."""
    if request.method == 'POST':
        sb_urls = request.form.getlist("SBurls")
        print("sb_urls:")
        pprint(sb_urls)
        projects = []
        for url in sb_urls:
            project_dict = {}
            proj = db.session.query(Project).filter(Project.url == url).first()
            if proj is None:
                print("---Error: Could not find project for {}".format(url))
                continue
            else:
                print("Found: {0}: {1}".format(proj.id, proj.name))
            fys = proj.fiscal_years
            if len(fys) > 1:
                project_dict['fy_id'] = []
                project_dict['casc_id'] = []
                for fy in fys:
                    project_dict['fy_id'].append(fy.id)
                    project_dict['casc_id'].append(fy.casc_id)
            else:
                fy = fys[0]
                project_dict['fy_id'] = fy.id
                project_dict['casc_id'] = fy.casc_id

            project_dict['proj_id'] = proj.id
            projects.append(project_dict)
            session["projects"] = projects
        return redirect(url_for('main.report'))

    return(render_template('projects.html', title="Select Projects to Report")),400


@bp.route('/report')
def report():
    """Gather appropriate report information and display."""

    excel_file = 'CASC Data Management Tracking for Projects - v2.xlsx'
    try:
        project_list = session["projects"]
    except KeyError:
        return render_template("error.html", message="Please select Fiscal Year First To Generate Report")
    # except TypeError:
    #     return render_template("error.html", message="Please Login ")
    projects = []
    workbook = None

    # Decide whether to load project tracking excel workbook
    if current_user.is_authenticated and current_user.access_level > 0:
        for project in project_list:
            casc_item = db.session.query(casc).get(project['casc_id'])
            if get_sheet_name(casc_item.name):
                # Load workbook
                try:
                    print('Opening {}...'.format(excel_file))
                    workbook = openpyxl.load_workbook(excel_file)
                    print('Successfully opened {}'.format(excel_file))
                except:
                    print('File error: {}'.format(excel_file))
                # No need to continue if workbook has just been loaded
                break
    
    class ReportItem(object):
        """Object to be passed to front-end for display in table and modal."""

        name = None
        id = None
        sb_id = None
        url = None
        obj_type = None
        data_in_project_GB = None
        num_of_files = None
        total_data_in_fy_GB = None
        timestamp = None
        dmp_status = None
        pi_list = []
        summary = None
        history = None
        item_breakdown = None
        potential_products = None
        products_received = []
        file_breakdown = []

        # Possibly necessary info:
        casc = None
        fiscal_year = None
        project = None
        item = None

        def __init__(i, obj_type, obj_db_id, fy_db_id, casc_db_id):
            """Initialize ReportItem class object.

            Arguments:
                obj_type -- (string) 'project', 'fiscal year', 'casc', 'item',
                            'sbfile', or 'problem item' to determine the type
                            of object being created.
                obj_db_id -- (int) the database id for the item being created.
                fy_db_id -- (int or list) the database id for the item's
                            fiscal year of concern.
                casc_db_id -- (int or list) the database id for the item's
                              casc year of concern.

            """

            sheet = {}

            if obj_type == 'project':
                i.obj_type = obj_type
                proj = db.session.query(Project).filter(
                    Project.id == obj_db_id).first()
                if proj == None:
                    raise Exception  # It has to be there somewhere...
                else:
                    i.name = proj.name
                    i.id = obj_db_id
                    i.sb_id = proj.sb_id
                    i.url = proj.url
                    # convert from MB -> GB
                    i.data_in_project_GB = proj.total_data / 1000
                    i.num_of_files = proj.files.count()
                    if fy_db_id is list:
                        i.fiscal_year = []
                        i.casc = []
                        i.total_data_in_fy_GB = []
                        for fy_id in fy_db_id:
                            fy = db.session.query(FiscalYear).get(fy_id)
                            i.fiscal_year.append(fy.name)
                            casc_model = db.session.query(casc).get(fy.casc_id)
                            i.casc.append(casc_model.name)
                            # convert from MB -> GB
                            i.total_data_in_fy_GB.append(
                                fy.total_data / 1000)

                    else:
                        fy = db.session.query(FiscalYear).get(fy_db_id)
                        i.fiscal_year = fy.name
                        casc_model = db.session.query(casc).get(casc_db_id)
                        i.casc = casc_model.name
                        # convert from MB -> GB
                        i.total_data_in_fy_GB = fy.total_data / 1000
                    i.timestamp = proj.timestamp
                    i.pi_list = []
                    for pi in proj.principal_investigators:
                        curr_pi = {'name': pi.name, 'email': pi.email}
                        i.pi_list.append(curr_pi)
                    i.summary = proj.summary
                    i.products_received = []
                    for item in proj.items:
                        curr_item = {'name': item.name, 'url': item.url}
                        i.products_received.append(curr_item)
                    # Things that depend on user access level:
                    if current_user.is_authenticated:
                        if current_user.access_level > 0:

                            # Parse excel sheet
                            sheet_name = get_sheet_name(i.casc)
                            if sheet_name:
                                values = []
                                for vals in workbook[sheet_name].values:
                                    if vals[0] is None:
                                        break
                                    values.append(vals)

                                sheet = parse_values(values)

    # ACTION ITEM: In a production app, you likely
                            #       want to save these credentials in a
                            #       persistent database instead.
                            session['credentials'] = credentials_to_dict(
                                credentials)
                            try:
                                # DMP Status
                                i.dmp_status = sheet[proj.sb_id]['DMP Status']
                                if i.dmp_status is None or i.dmp_status.isspace() or i.dmp_status == "":

                                    i.dmp_status = "No DMP status provided"
                                # History
                                i.history = sheet[proj.sb_id]['History']
                                if i.history is None or i.history.isspace() or i.history == "":
                                    i.history = "No data steward history provided"
                                # Potential Products
                                i.potential_products = sheet[proj.sb_id]['Expected Products']
                                if i.potential_products is None or i.potential_products.isspace() or i.potential_products == "":
                                    i.potential_products = "No data potential products provided"
                            except KeyError:
                                i.dmp_status = "Project not currently tracked by Data Steward"
                                i.history = "Project not currently tracked by Data Steward"
                                i.potential_products = "Project not currently tracked by Data Steward"

                        else:
                            i.dmp_status = "Please email administrators at"\
                                + " {} to receive access privileges to view "\
                                .format(current_app.config['ADMINS'][0])\
                                + "this content."
                            i.history = "Please email administrators at"\
                                + " {} to receive access privileges to view "\
                                .format(current_app.config['ADMINS'][0])\
                                + "this content."
                            i.potential_products = "Please email "\
                                + "administrators at {} to receive access "\
                                .format(current_app.config['ADMINS'][0])\
                                + "privileges to view this content."
                    else:
                        i.dmp_status = "Please login to view this content."
                        i.history = "Please login to view this content."
                        i.potential_products = "Please login to view this content."
                    i.file_breakdown = []
                    proj_file_list = []
                    for sbfile in proj.files:
                        proj_file_list.append(sbfile.id)
                    if len(proj_file_list) > 0:
                        file_breakdown_list = db.session.query(
                            SbFile.content_type, db.func.count(
                                SbFile.content_type)).group_by(
                                    SbFile.content_type).filter(
                                        SbFile.id.in_(proj_file_list[:999])).all()  # sqlalchemy max query items is 999
                        proj_file_list[:] = []
                        for _tuple in file_breakdown_list:
                            temp_dict = {}
                            temp_dict['label'] = _tuple[0]
                            temp_dict['count'] = _tuple[1]
                            proj_file_list.append(temp_dict)
                        i.file_breakdown = sorted(
                            proj_file_list,
                            key=lambda k: k['count'],
                            reverse=True)

            elif obj_type == 'fiscal year':
                pass  # We don't do anything with fiscal year objects on the
                # front-end yet.
            elif obj_type == 'casc':
                pass  # We don't do anything with fiscal year objects on the
                # front-end yet.
            elif obj_type == 'item':
                pass  # We don't do anything with fiscal year objects on the
                # front-end yet.
            elif obj_type == 'sbfile':
                pass  # We don't do anything with fiscal year objects on the
                # front-end yet.
            elif obj_type == 'problem item':
                pass  # We don't do anything with fiscal year objects on the
                # front-end yet.
    for project in project_list:
        new_obj = ReportItem(
            'project', project['proj_id'], project['fy_id'], project['casc_id'])
        projects.append(new_obj.__dict__)

    return render_template("report.html", projects=projects)

@bp.route('/verticalbar')
def verticalbar():
    """Gather appropriate report information and display."""

    excel_file = 'CASC Data Management Tracking for Projects - v2.xlsx'
    project_list = session["projects"]
    projects = []
    workbook = None

    # Decide whether to load project tracking excel workbook
    if current_user.is_authenticated and current_user.access_level > 0:
        for project in project_list:
            casc_item = db.session.query(casc).get(project['casc_id'])
            if get_sheet_name(casc_item.name):
                # Load workbook
                try:
                    print('Opening {}...'.format(excel_file))
                    workbook = openpyxl.load_workbook(excel_file)
                    print('Successfully opened {}'.format(excel_file))
                except:
                    print('File error: {}'.format(excel_file))
                # No need to continue if workbook has just been loaded
                break

    class ReportItem(object):
        """Object to be passed to front-end for display in table and modal."""

        name = None
        id = None
        sb_id = None
        url = None
        obj_type = None
        data_in_project_GB = None
        num_of_files = None
        total_data_in_fy_GB = None
        timestamp = None
        dmp_status = None
        pi_list = []
        summary = None
        history = None
        item_breakdown = None
        potential_products = None
        products_received = []
        file_breakdown = []

        # Possibly necessary info:
        casc = None
        fiscal_year = None
        project = None
        item = None

        def __init__(i, obj_type, obj_db_id, fy_db_id, casc_db_id):
            """Initialize ReportItem class object.

            Arguments:
                obj_type -- (string) 'project', 'fiscal year', 'casc', 'item',
                            'sbfile', or 'problem item' to determine the type
                            of object being created.
                obj_db_id -- (int) the database id for the item being created.
                fy_db_id -- (int or list) the database id for the item's
                            fiscal year of concern.
                casc_db_id -- (int or list) the database id for the item's
                              casc year of concern.

            """

            sheet = {}

            if obj_type == 'project':
                i.obj_type = obj_type
                proj = db.session.query(Project).filter(
                    Project.id == obj_db_id).first()
                if proj == None:
                    raise Exception  # It has to be there somewhere...
                else:
                    i.name = proj.name
                    i.id = obj_db_id
                    i.sb_id = proj.sb_id
                    i.url = proj.url
                    # convert from MB -> GB
                    i.data_in_project_GB = proj.total_data / 1000
                    i.num_of_files = proj.files.count()
                    if fy_db_id is list:
                        i.fiscal_year = []
                        i.casc = []
                        i.total_data_in_fy_GB = []
                        for fy_id in fy_db_id:
                            fy = db.session.query(FiscalYear).get(fy_id)
                            i.fiscal_year.append(fy.name)
                            casc_model = db.session.query(casc).get(fy.casc_id)
                            i.casc.append(casc_model.name)
                            # convert from MB -> GB
                            i.total_data_in_fy_GB.append(
                                fy.total_data / 1000)

                    else:
                        fy = db.session.query(FiscalYear).get(fy_db_id)
                        i.fiscal_year = fy.name
                        casc_model = db.session.query(casc).get(casc_db_id)
                        i.casc = casc_model.name
                        # convert from MB -> GB
                        i.total_data_in_fy_GB = fy.total_data / 1000
                    i.timestamp = proj.timestamp
                    i.pi_list = []
                    for pi in proj.principal_investigators:
                        curr_pi = {'name': pi.name, 'email': pi.email}
                        i.pi_list.append(curr_pi)
                    i.summary = proj.summary
                    i.products_received = []
                    for item in proj.items:
                        curr_item = {'name': item.name, 'url': item.url}
                        i.products_received.append(curr_item)
                    # Things that depend on user access level:
                    if current_user.is_authenticated:
                        if current_user.access_level > 0:

                            # Parse excel sheet
                            sheet_name = get_sheet_name(i.casc)
                            if sheet_name:
                                values = []
                                for vals in workbook[sheet_name].values:
                                    if vals[0] is None:
                                        break
                                    values.append(vals)

                                sheet = parse_values(values)

    # ACTION ITEM: In a production app, you likely
                            #       want to save these credentials in a
                            #       persistent database instead.
                            session['credentials'] = credentials_to_dict(
                                credentials)
                            try:
                                # DMP Status
                                i.dmp_status = sheet[proj.sb_id]['DMP Status']
                                if i.dmp_status is None or i.dmp_status.isspace() or i.dmp_status == "":

                                    i.dmp_status = "No DMP status provided"
                                # History
                                i.history = sheet[proj.sb_id]['History']
                                if i.history is None or i.history.isspace() or i.history == "":
                                    i.history = "No data steward history provided"
                                # Potential Products
                                i.potential_products = sheet[proj.sb_id]['Expected Products']
                                if i.potential_products is None or i.potential_products.isspace() or i.potential_products == "":
                                    i.potential_products = "No data potential products provided"
                            except KeyError:
                                i.dmp_status = "Project not currently tracked by Data Steward"
                                i.history = "Project not currently tracked by Data Steward"
                                i.potential_products = "Project not currently tracked by Data Steward"

                        else:
                            i.dmp_status = "Please email administrators at"\
                                + " {} to receive access privileges to view "\
                                .format(current_app.config['ADMINS'][0])\
                                + "this content."
                            i.history = "Please email administrators at"\
                                + " {} to receive access privileges to view "\
                                .format(current_app.config['ADMINS'][0])\
                                + "this content."
                            i.potential_products = "Please email "\
                                + "administrators at {} to receive access "\
                                .format(current_app.config['ADMINS'][0])\
                                + "privileges to view this content."
                    else:
                        i.dmp_status = "Please login to view this content."
                        i.history = "Please login to view this content."
                        i.potential_products = "Please login to view this content."
                    i.file_breakdown = []
                    proj_file_list = []
                    for sbfile in proj.files:
                        proj_file_list.append(sbfile.id)
                    if len(proj_file_list) > 0:
                        file_breakdown_list = db.session.query(
                            SbFile.content_type, db.func.count(
                                SbFile.content_type)).group_by(
                                    SbFile.content_type).filter(
                                        SbFile.id.in_(proj_file_list[:999])).all()  # sqlalchemy max query items is 999
                        proj_file_list[:] = []
                        for _tuple in file_breakdown_list:
                            temp_dict = {}
                            temp_dict['label'] = _tuple[0]
                            temp_dict['count'] = _tuple[1]
                            proj_file_list.append(temp_dict)
                        i.file_breakdown = sorted(
                            proj_file_list,
                            key=lambda k: k['count'],
                            reverse=True)

            elif obj_type == 'fiscal year':
                pass  # We don't do anything with fiscal year objects on the
                # front-end yet.
            elif obj_type == 'casc':
                pass  # We don't do anything with fiscal year objects on the
                # front-end yet.
            elif obj_type == 'item':
                pass  # We don't do anything with fiscal year objects on the
                # front-end yet.
            elif obj_type == 'sbfile':
                pass  # We don't do anything with fiscal year objects on the
                # front-end yet.
            elif obj_type == 'problem item':
                pass  # We don't do anything with fiscal year objects on the
                # front-end yet.

    for project in project_list:
        new_obj = ReportItem(
            'project', project['proj_id'], project['fy_id'], project['casc_id'])
        projects.append(new_obj.__dict__)

    return render_template("verticalbar.html", projects=projects)


@bp.route('/horizontalbar')
def horizontalbar():
    """Gather appropriate report information and display."""

    excel_file = 'CASC Data Management Tracking for Projects - v2.xlsx'
    project_list = session["projects"]
    projects = []
    workbook = None

    # Decide whether to load project tracking excel workbook
    if current_user.is_authenticated and current_user.access_level > 0:
        for project in project_list:
            casc_item = db.session.query(casc).get(project['casc_id'])
            if get_sheet_name(casc_item.name):
                # Load workbook
                try:
                    print('Opening {}...'.format(excel_file))
                    workbook = openpyxl.load_workbook(excel_file)
                    print('Successfully opened {}'.format(excel_file))
                except:
                    print('File error: {}'.format(excel_file))
                # No need to continue if workbook has just been loaded
                break

    class ReportItem(object):
        """Object to be passed to front-end for display in table and modal."""

        name = None
        id = None
        sb_id = None
        url = None
        obj_type = None
        data_in_project_GB = None
        num_of_files = None
        total_data_in_fy_GB = None
        timestamp = None
        dmp_status = None
        pi_list = []
        summary = None
        history = None
        item_breakdown = None
        potential_products = None
        products_received = []
        file_breakdown = []

        # Possibly necessary info:
        casc = None
        fiscal_year = None
        project = None
        item = None

        def __init__(i, obj_type, obj_db_id, fy_db_id, casc_db_id):
            """Initialize ReportItem class object.

            Arguments:
                obj_type -- (string) 'project', 'fiscal year', 'casc', 'item',
                            'sbfile', or 'problem item' to determine the type
                            of object being created.
                obj_db_id -- (int) the database id for the item being created.
                fy_db_id -- (int or list) the database id for the item's
                            fiscal year of concern.
                casc_db_id -- (int or list) the database id for the item's
                              casc year of concern.

            """

            sheet = {}

            if obj_type == 'project':
                i.obj_type = obj_type
                proj = db.session.query(Project).filter(Project.id == obj_db_id).first()
                if proj == None:
                    raise Exception  # It has to be there somewhere...
                else:
                    i.name = proj.name
                    i.id = obj_db_id
                    i.sb_id = proj.sb_id
                    i.url = proj.url
                    # convert from MB -> GB
                    i.data_in_project_GB = proj.total_data / 1000
                    i.num_of_files = proj.files.count()
                    if fy_db_id is list:
                        i.fiscal_year = []
                        i.casc = []
                        i.total_data_in_fy_GB = []
                        for fy_id in fy_db_id:
                            fy = db.session.query(FiscalYear).get(fy_id)
                            i.fiscal_year.append(fy.name)
                            casc_model = db.session.query(casc).get(fy.casc_id)
                            i.casc.append(casc_model.name)
                            # convert from MB -> GB
                            i.total_data_in_fy_GB.append(fy.total_data / 1000)

                    else:
                        fy = db.session.query(FiscalYear).get(fy_db_id)
                        i.fiscal_year = fy.name
                        casc_model = db.session.query(casc).get(casc_db_id)
                        i.casc = casc_model.name
                        # convert from MB -> GB
                        i.total_data_in_fy_GB = fy.total_data / 1000
                    i.timestamp = proj.timestamp
                    i.pi_list = []
                    for pi in proj.principal_investigators:
                        curr_pi = {'name': pi.name, 'email': pi.email}
                        i.pi_list.append(curr_pi)
                    i.summary = proj.summary
                    i.products_received = []
                    for item in proj.items:
                        curr_item = {'name': item.name, 'url': item.url}
                        i.products_received.append(curr_item)
                    # Things that depend on user access level:
                    if current_user.is_authenticated:
                        if current_user.access_level > 0:

                            # Parse excel sheet
                            sheet_name = get_sheet_name(i.casc)
                            if sheet_name:
                                values = []
                                try:
                                    for vals in workbook[sheet_name].values:
                                        if vals[0] is None:
                                            break
                                        values.append(vals)

                                    sheet = parse_values(values)
                                except:
                                    pass

    # ACTION ITEM: In a production app, you likely
                            #       want to save these credentials in a
                            #       persistent database instead.
                            session['credentials'] = credentials_to_dict(credentials)
                            try:
                                try:
                                    # DMP Status
                                    i.dmp_status = sheet[proj.sb_id]['DMP Status']
                                    if i.dmp_status is None or i.dmp_status.isspace() or i.dmp_status == "":

                                        i.dmp_status = "No DMP status provided"
                                    # History
                                    i.history = sheet[proj.sb_id]['History']
                                    if i.history is None or i.history.isspace() or i.history == "":
                                        i.history = "No data steward history provided"
                                    # Potential Products
                                    i.potential_products = sheet[proj.sb_id]['Expected Products']
                                    if i.potential_products is None or i.potential_products.isspace() or i.potential_products == "":
                                        i.potential_products = "No data potential products provided"
                                except KeyError:
                                    i.dmp_status = "Project not currently tracked by Data Steward"
                                    i.history = "Project not currently tracked by Data Steward"
                                    i.potential_products = "Project not currently tracked by Data Steward"
                            except:
                                pass
                        else:
                            i.dmp_status = "Please email administrators at"\
                                + " {} to receive access privileges to view "\
                                .format(current_app.config['ADMINS'][0])\
                                + "this content."
                            i.history = "Please email administrators at"\
                                + " {} to receive access privileges to view "\
                                .format(current_app.config['ADMINS'][0])\
                                + "this content."
                            i.potential_products = "Please email "\
                                + "administrators at {} to receive access "\
                                .format(current_app.config['ADMINS'][0])\
                                + "privileges to view this content."
                    else:
                        i.dmp_status = "Please login to view this content."
                        i.history = "Please login to view this content."
                        i.potential_products = "Please login to view this content."
                    i.file_breakdown = []
                    proj_file_list = []
                    for sbfile in proj.files:
                        proj_file_list.append(sbfile.id)
                    if len(proj_file_list) > 0:
                        file_breakdown_list = db.session.query(
                            SbFile.content_type, db.func.count(
                                SbFile.content_type)).group_by(
                                    SbFile.content_type).filter(
                                        SbFile.id.in_(proj_file_list[:999])).all()  # sqlalchemy max query items is 999
                        proj_file_list[:] = []
                        for _tuple in file_breakdown_list:
                            temp_dict = {}
                            temp_dict['label'] = _tuple[0]
                            temp_dict['count'] = _tuple[1]
                            proj_file_list.append(temp_dict)
                        i.file_breakdown = sorted(
                            proj_file_list,
                            key=lambda k: k['count'],
                            reverse=True)

            elif obj_type == 'fiscal year':
                pass  # We don't do anything with fiscal year objects on the
                # front-end yet.
            elif obj_type == 'casc':
                pass  # We don't do anything with fiscal year objects on the
                # front-end yet.
            elif obj_type == 'item':
                pass  # We don't do anything with fiscal year objects on the
                # front-end yet.
            elif obj_type == 'sbfile':
                pass  # We don't do anything with fiscal year objects on the
                # front-end yet.
            elif obj_type == 'problem item':
                pass  # We don't do anything with fiscal year objects on the
                # front-end yet.

    for project in project_list:
        new_obj = ReportItem(
            'project', project['proj_id'], project['fy_id'], project['casc_id'])
        projects.append(new_obj.__dict__)

    return render_template("horizontalbar.html", projects=projects)



@bp.route('/treemap')
def treemap():

    """Gather appropriate report information and display."""

    excel_file = 'CASC Data Management Tracking for Projects - v2.xlsx'
    project_list = session["projects"]

    projects = []
    workbook = None

    # Decide whether to load project tracking excel workbook
    if current_user.is_authenticated and current_user.access_level > 0:
        for project in project_list:
            casc_item = db.session.query(casc).get(project['casc_id'])
            if get_sheet_name(casc_item.name):
                # Load workbook
                try:
                    print('Opening {}...'.format(excel_file))
                    workbook = openpyxl.load_workbook(excel_file)
                    print('Successfully opened {}'.format(excel_file))
                except:
                    print('File error: {}'.format(excel_file))
                # No need to continue if workbook has just been loaded
                break

    class ReportItem(object):
        """Object to be passed to front-end for display in table and modal."""

        name = None
        id = None
        sb_id = None
        url = None
        obj_type = None
        data_in_project_GB = None
        num_of_files = None
        total_data_in_fy_GB = None
        timestamp = None
        dmp_status = None
        pi_list = []
        summary = None
        history = None
        item_breakdown = None
        potential_products = None
        products_received = []
        file_breakdown = []

        # Possibly necessary info:
        casc = None
        fiscal_year = None
        project = None
        item = None

        def __init__(i, obj_type, obj_db_id, fy_db_id, casc_db_id):
            """Initialize ReportItem class object.

            Arguments:
                obj_type -- (string) 'project', 'fiscal year', 'casc', 'item',
                            'sbfile', or 'problem item' to determine the type
                            of object being created.
                obj_db_id -- (int) the database id for the item being created.
                fy_db_id -- (int or list) the database id for the item's
                            fiscal year of concern.
                casc_db_id -- (int or list) the database id for the item's
                              casc year of concern.

            """

            sheet = {}

            if obj_type == 'project':
                i.obj_type = obj_type
                proj = db.session.query(Project).filter(
                    Project.id == obj_db_id).first()
                if proj == None:
                    raise Exception  # It has to be there somewhere...
                else:
                    i.name = proj.name
                    i.id = obj_db_id
                    i.sb_id = proj.sb_id
                    i.url = proj.url
                    # convert from MB -> GB
                    i.data_in_project_GB = proj.total_data / 1000
                    i.num_of_files = proj.files.count()
                    if fy_db_id is list:
                        i.fiscal_year = []
                        i.casc = []
                        i.total_data_in_fy_GB = []
                        for fy_id in fy_db_id:
                            fy = db.session.query(FiscalYear).get(fy_id)
                            i.fiscal_year.append(fy.name)
                            casc_model = db.session.query(casc).get(fy.casc_id)
                            i.casc.append(casc_model.name)
                            # convert from MB -> GB
                            i.total_data_in_fy_GB.append(
                                fy.total_data / 1000)

                    else:
                        fy = db.session.query(FiscalYear).get(fy_db_id)
                        i.fiscal_year = fy.name
                        casc_model = db.session.query(casc).get(casc_db_id)
                        i.casc = casc_model.name
                        # convert from MB -> GB
                        i.total_data_in_fy_GB = fy.total_data / 1000
                    i.timestamp = proj.timestamp
                    i.pi_list = []
                    for pi in proj.principal_investigators:
                        curr_pi = {'name': pi.name, 'email': pi.email}
                        i.pi_list.append(curr_pi)
                    i.summary = proj.summary
                    i.products_received = []
                    for item in proj.items:
                        curr_item = {'name': item.name, 'url': item.url}
                        i.products_received.append(curr_item)
                    # Things that depend on user access level:
                    if current_user.is_authenticated:
                        if current_user.access_level > 0:

                            # Parse excel sheet
                            sheet_name = get_sheet_name(i.casc)
                            if sheet_name:
                                values = []
                                for vals in workbook[sheet_name].values:
                                    if vals[0] is None:
                                        break
                                    values.append(vals)

                                sheet = parse_values(values)

    # ACTION ITEM: In a production app, you likely
                            #       want to save these credentials in a
                            #       persistent database instead.
                            session['credentials'] = credentials_to_dict(
                                credentials)
                            try:
                                # DMP Status
                                i.dmp_status = sheet[proj.sb_id]['DMP Status']
                                if i.dmp_status is None or i.dmp_status.isspace() or i.dmp_status == "":

                                    i.dmp_status = "No DMP status provided"
                                # History
                                i.history = sheet[proj.sb_id]['History']
                                if i.history is None or i.history.isspace() or i.history == "":
                                    i.history = "No data steward history provided"
                                # Potential Products
                                i.potential_products = sheet[proj.sb_id]['Expected Products']
                                if i.potential_products is None or i.potential_products.isspace() or i.potential_products == "":
                                    i.potential_products = "No data potential products provided"
                            except KeyError:
                                i.dmp_status = "Project not currently tracked by Data Steward"
                                i.history = "Project not currently tracked by Data Steward"
                                i.potential_products = "Project not currently tracked by Data Steward"

                        else:
                            i.dmp_status = "Please email administrators at"\
                                + " {} to receive access privileges to view "\
                                .format(current_app.config['ADMINS'][0])\
                                + "this content."
                            i.history = "Please email administrators at"\
                                + " {} to receive access privileges to view "\
                                .format(current_app.config['ADMINS'][0])\
                                + "this content."
                            i.potential_products = "Please email "\
                                + "administrators at {} to receive access "\
                                .format(current_app.config['ADMINS'][0])\
                                + "privileges to view this content."
                    else:
                        i.dmp_status = "Please login to view this content."
                        i.history = "Please login to view this content."
                        i.potential_products = "Please login to view this content."
                    i.file_breakdown = []
                    proj_file_list = []
                    for sbfile in proj.files:
                        proj_file_list.append(sbfile.id)
                    if len(proj_file_list) > 0:
                        file_breakdown_list = db.session.query(
                            SbFile.content_type, db.func.count(
                                SbFile.content_type)).group_by(
                                    SbFile.content_type).filter(
                                        SbFile.id.in_(proj_file_list[:999])).all()  # sqlalchemy max query items is 999
                        proj_file_list[:] = []
                        for _tuple in file_breakdown_list:
                            temp_dict = {}
                            temp_dict['label'] = _tuple[0]
                            temp_dict['count'] = _tuple[1]
                            proj_file_list.append(temp_dict)
                        i.file_breakdown = sorted(
                            proj_file_list,
                            key=lambda k: k['count'],
                            reverse=True)

            elif obj_type == 'fiscal year':
                pass  # We don't do anything with fiscal year objects on the
                # front-end yet.
            elif obj_type == 'casc':
                pass  # We don't do anything with fiscal year objects on the
                # front-end yet.
            elif obj_type == 'item':
                pass  # We don't do anything with fiscal year objects on the
                # front-end yet.
            elif obj_type == 'sbfile':
                pass  # We don't do anything with fiscal year objects on the
                # front-end yet.
            elif obj_type == 'problem item':
                pass  # We don't do anything with fiscal year objects on the
                # front-end yet.

    for project in project_list:
        new_obj = ReportItem(
            'project', project['proj_id'], project['fy_id'], project['casc_id'])
        projects.append(new_obj.__dict__)

    return render_template("treemap.html", projects=projects)

@bp.route('/user/<username>')
@login_required
def user(username):
    """Load user and render user.html template if found, else 404."""
    # Change to lowercase to make case insensitive
    user = User.query.filter_by(username=username.lower()).first_or_404()
    admin_email = current_app.config['ADMINS'][0]
    return render_template('user.html', user=user, adminEmail=admin_email)


@bp.route('/edit_profile', methods=['GET', 'POST'])
@login_required
def edit_profile():
    """Define form for editing a profile."""
    form = EditProfileForm(current_user.username)
    if form.validate_on_submit():

        if form.username.data:
            current_user.username = str(form.username.data).lower()
        if form.about.data:
            current_user.about = str(form.about.data)
        if form.email.data:
            current_user.email = str(form.email.data)
        if form.password.data:
            # current_user.password = str(form.password.data)
            user = current_user
            user.set_password(form.password.data)
            db.session.add(user)
        db.session.commit()
        return redirect(url_for('main.user', username=current_user.username)),400
    elif request.method == 'GET':
        form.username.data = current_user.username
        form.about.data = current_user.about
        form.email.data = current_user.email

    return render_template(
        'edit_profile.html', title='Edit Profile', form=form),400



@bp.route('/search', methods=['GET', 'POST'])
@login_required
def search():
    current_user.search_form = SearchForm()
    d= str(current_user.search_form.data['search'])
    if(len(d)==0):
        message=["Please Enter The Keyword To Search"]
        return render_template('error.html', message = message, length=len(d)),400

    courses = MasterDetails.query.filter((MasterDetails.projectTitle.like('%'+d+'%')) | (MasterDetails.PI.like('%'+d+'%'))) .all()
    length=len(courses)
    
    # Adding Required Details to userData
    userdata=[]
    def add_user(user):
        userdata.append(user)
    for i in courses:
        user = {}
        user["name"]=i.projectTitle
        user["casc"]=i.casc
        user["Fy"]=str(i.fy)
        user["ctitle"]=i.title
        user["sbId"]=i.sb_id
        user["summary"]=i.summary
        add_user(user)
    return render_template('searchTree.html',query=d, courses = courses, length= length, userdata=userdata)

@bp.route('/searchBar/<query>', methods=['GET', 'POST'])
@login_required
def searchBar(query):

    current_user.search_form = SearchForm()
    d=query
    if(len(d)==0):
        userdata=["Please Enter The Keyword To Search"]
        return render_template('search_results.html', userdata = userdata, length=len(d))

    courses = MasterDetails.query.filter((MasterDetails.projectTitle.like('%'+d+'%')) | (MasterDetails.PI.like('%'+d+'%'))) .all()
    length=len(courses)
    # Adding Required Details to userData
    userdata=[]
    def add_user(user):
        userdata.append(user)
    for i in courses:
        user = {}
        user["name"]=i.projectTitle
        user["casc"]=i.casc
        user["Fy"]=str(i.fy)
        user["ctitle"]=i.title
        user["size"]=i.projectSize
        add_user(user)
    return render_template('search_results.html', query=d, courses=courses, length=length, userdata=userdata)

@bp.route('/searchBack/<query>', methods=['GET', 'POST'])
@login_required
def searchBack(query):

    current_user.search_form = SearchForm()
    d=query
    if(len(d)==0):
        userdata=["Please Enter The Keyword To Search"]
        return render_template('search_results.html', userdata = userdata, length=len(d))

    courses = MasterDetails.query.filter((MasterDetails.projectTitle.like('%'+d+'%')) | (MasterDetails.PI.like('%'+d+'%'))) .all()
    length=len(courses)
    # Adding Required Details to userData
    userdata=[]
    def add_user(user):
        userdata.append(user)
    for i in courses:
        user = {}
        user["name"]=i.projectTitle
        user["casc"]=i.casc
        user["Fy"]=str(i.fy)
        user["ctitle"]=i.title
        user["size"]=i.projectSize
        
        add_user(user)
    return render_template('searchTree.html', query=d, courses=courses, length=length, userdata=userdata)

@bp.route('/searchTable/<query>', methods=['GET', 'POST'])
@login_required
def searchTable(query):
    current_user.search_form = SearchForm()
    d=query
    if(len(d)==0):
        userdata=["Please Enter The Keyword To Search"]
        return render_template('search_results.html', userdata = userdata, length=len(d))

    courses = MasterDetails.query.filter((MasterDetails.projectTitle.like('%'+d+'%')) | (MasterDetails.PI.like('%'+d+'%'))) .all()
    length=len(courses)
    userdata=[]
    def add_user(user):
        userdata.append(user)
    sheet = {}
    for i in courses:
        user = {}
        user["name"]=i.projectTitle
        user['sb_id']=i.sb_id
        user["casc"]=i.casc
        user["Fy"]=str(i.fy)
        user["ctitle"]=i.title
        user["size"]=i.projectSize
        user['pi_list']=i.PI
        user['summary']=i.summary
        user['url']=i.url
        user['ctitle']=i.title
        user['xml']=i.xml_urls
        if(i.xml_urls==''):
            # print("no data")
            user['xml']="Metadata Unavailable for this DataItem"
            user['error']="No Validations"
            user['curl']="No Xml Link Found"
        else:
            doc = etree.parse(path_to_static + 'FGDC-BDP/fgdc-std-001.1-1999.xsd')
            schema = etree.XMLSchema(doc)

            #parse the url and convert to xml file 
            url1=str(i.xml_urls)
            URL = url1.split(',')[0]
            user['curl']=URL
            # print(URL)
            try:
                response = requests.get(URL)
                with open(path_to_static + 'feed.xml', 'wb') as file:
                    file.write(response.content)
                # Schema to be validated.
                custom = etree.parse(path_to_static + 'feed.xml')

                # Validate Schema
                user['xml']=schema.validate(custom)
                # print(schema.validate(custom))

                def get_project(error):
                    # return error.path.split('/')[-1]
                    return error.message.split(':')[0].split(" ")[1].strip()


                # If errors, we will find it in schema.error_log
                user['error']=[]
                for error in schema.error_log:
                    # Mutiple attribute available in error
                    # 'column', 'domain', 'domain_name', 'filename', 'level', 'level_name',
                    # 'line', 'message', 'path', 'type', 'type_name'
                    error1=str(error.message),"Error in Line Number: "+str(error.line)
                    # print('ErrorMessage',error.message)
                    user['error'].append(error1)           
                result = {}
                for error1 in schema.error_log:
                    project = get_project(error1)
                    if project not in result:
                        result[project] = 0

                    result[project] += 1
                    # print(error.message)
                # print(result)
                user['countError']=str(result)
            except:
                user['cxml']="URL Associated with this Data Item is not working"
                # print("url failed")

        if current_user.is_authenticated:
            if current_user.access_level > 0:

                # Parse excel sheet
                sheet_name = get_sheet_name(i.casc)
                if sheet_name:
                    values = []
                    for vals in workbook[sheet_name].values:
                        if vals[0] is None:
                            break
                        values.append(vals)

                    sheet = parse_values(values)

                try:
                    # DMP Status
                    i.dmp_status = sheet[i.sb_id]['DMP Status']
                    if i.dmp_status is None or i.dmp_status.isspace() or i.dmp_status == "":
                        i.dmp_status = "No DMP status provided"
                    user["dmp_status"]= i.dmp_status
                    # History
                    i.history = sheet[i.sb_id]['History']
                    if i.history is None or i.history.isspace() or i.history == "":
                        i.history = "No data steward history provided"
                    user['history']=i.history
                    # Potential Products
                    i.potential_products = sheet[i.sb_id]['Expected Products']
                    if i.potential_products is None or i.potential_products.isspace() or i.potential_products == "":
                        i.potential_products = "No data potential products provided"
                    user['potential_products']= i.potential_products
                except KeyError:
                    i.dmp_status = "Project not currently tracked by Data Steward"
                    i.history = "Project not currently tracked by Data Steward"
                    i.potential_products = "Project not currently tracked by Data Steward"
            else:
                i.dmp_status = "Please email administrators at"\
                    + " {} to receive access privileges to view "\
                    .format(current_app.config['ADMINS'][0])\
                    + "this content."
                user["dmp_status"]= i.dmp_status
                i.history = "Please email administrators at"\
                    + " {} to receive access privileges to view "\
                    .format(current_app.config['ADMINS'][0])\
                    + "this content."
                user['history']=i.history
                i.potential_products = "Please email "\
                    + "administrators at {} to receive access "\
                    .format(current_app.config['ADMINS'][0])\
                    + "privileges to view this content."
                user['potential_products']= i.potential_products
        else:
            i.dmp_status = "Please login to view this content."
            i.history = "Please login to view this content."
            i.potential_products = "Please login to view this content."
        i.file_breakdown = []
        add_user(user)
    return render_template('searchTableChart.html', query=d, courses=courses, length=length, userdata=userdata)