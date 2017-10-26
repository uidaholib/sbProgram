from pandas import DataFrame
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

from pprint import pprint

import requests
import json
import pysb

sb = pysb.SbSession()

def mainExcel(ChosenFiscalYear, L1Project, L2DataInProject, L3NumSbItemsFound,
    L5RunningDataTotal, L6NestedData, L7MissingData,
    L8MissingDataURL, L9Exceptions, L10Exceptions_IDs, CouldNotFindFiles_id,
    DataHostedElsewhere_id, PossiblePermissionsIssues_id):
    #try:
    PossiblePermissionsIssuesURL = []

    for i in PossiblePermissionsIssues_id:
        json = sb.get_item(i)
        PossiblePermissionsIssuesURL.append(json['link']['url'])

    wb = Workbook()
    ws = wb.active
    if L8MissingDataURL == []:
        pass
    elif L8MissingDataURL != []:
        ws_missing = wb.create_sheet("Missing", 1) #Inserts a new sheet, named "Missing" in the second position
        #ws_missing.sheet_properties.tabColor = "1072BA" #Makes the tab for the sheet red so it draws attention.
    if PossiblePermissionsIssuesURL == []:
        pass
    elif PossiblePermissionsIssuesURL != []:
        ws_exceptions = wb.create_sheet("Exceptions", 2) #Inserts a new sheet, named "Exceptions" in the third position
        #ws_exceptions.sheet_properties.tabColor = "1072BA" #Makes the tab for the sheet red so it draws attention.


    df = DataFrame({'Project': L1Project, 'Data in project (GB)': L2DataInProject,
                        'Number of SB Items Found': L3NumSbItemsFound, 'Running Data Total (GB)':
                        L5RunningDataTotal, 'Nested Data?': L6NestedData,
                        })
                        #include these eventually: 'Missing Data?': L7MissingData, 'Exceptions/Permissions Issues': L9Exceptions

    dfOrdered = df[['Project', 'Data in project (GB)', 'Number of SB Items Found',
                'Running Data Total (GB)', 'Nested Data?',
                ]]
                #include these eventually: 'Missing Data?', 'Exceptions/Permissions Issues'

    df_missing = DataFrame({'Missing Data URL': L8MissingDataURL})


    df_exceptions = DataFrame({'Exception/Permission Issue URL': PossiblePermissionsIssuesURL}) #include 'Exception ID': L10Exceptions_IDs, later
    df_exceptionsOrdered = df_exceptions[['Exception/Permission Issue URL']] #include 'Exception ID' later


    for r in dataframe_to_rows(dfOrdered, index=False, header=True):
        ws.append(r)

    for cell in ws[1]:
        cell.style = 'Pandas'

    if L8MissingDataURL == []:
        pass
    elif L8MissingDataURL != []:
        for r in dataframe_to_rows(df_missing, index=False, header=True):
            ws_missing.append(r)

        for cell in ws_missing[1]:
            cell.style = 'Pandas'

    if PossiblePermissionsIssuesURL == []:
        pass
    elif PossiblePermissionsIssuesURL != []:
        for r in dataframe_to_rows(df_exceptionsOrdered, index=False, header=True):
            ws_exceptions.append(r)

        for cell in ws_exceptions[1]:
            cell.style = 'Pandas'

    print(df)
    print('''

    ---------Let\'s try reordering that...

    ''')
    print(dfOrdered)

    print ('''
    Would you like to save this?)
    (Y / N)''')
    answer = input('> ')
    if 'y' in answer or 'Y' in answer or "yes" in answer or "Yes" in answer:
        saveExcel(ChosenFiscalYear, wb)
    elif 'no' in answer or 'No' in answer or 'n' in answer or 'N' in answer:
        print('''
        Ok, we won't save it.''')
        import specialtyTasks_working3
        specialtyTasks_working3.nowWhat()
#    except (ValueError, Exception) as e:
#        print('''
#    ----------------------------WARNING: Something went wrong in the function "mainExcel" in ExcelPrint.py.''')
#        exit()


def saveExcel(ChosenFiscalYear, wb):
    print('''
    Would you like to name it something other than "'''+str(ChosenFiscalYear)+
        ''' Data Metrics.xlsx"?''')
    answer = input('> ')
    if 'y' in answer or 'Y' in answer or "yes" in answer or "Yes" in answer:
        print('''
    What name would you like to give the Excel file?''')
        name = input('> ')
        wb.save(str(name)+".xlsx")
        print('''
    Workbook saved as "'''+str(name)+'''.xlsx" in the working directory.''')
        import specialtyTasks_working3
        specialtyTasks_working3.nowWhat()
    elif 'no' in answer or 'No' in answer or 'n' in answer or 'N' in answer:
        wb.save(str(ChosenFiscalYear)+" Data Metrics.xlsx")
        print('''
    Workbook saved as "'''+str(ChosenFiscalYear)+''' Data Metrics.xlsx" in the working directory.''')
        import specialtyTasks_working3
        specialtyTasks_working3.nowWhat()
