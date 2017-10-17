import gl

from flask import Flask, render_template, redirect, \
    url_for, request, session, flash, jsonify, send_from_directory, send_file
import os
import pandas
from pandas import DataFrame
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

from pprint import pprint

import requests
import json
import pysb

import io  # For the creation of download file in-memory


sb = pysb.SbSession()

class sbItem(object):
    
    def __init__(self):
        self.ID = "No information provided"
        self.ObjectType = "No information provided"
        self.name = "No information provided"
        self.FY = "No information provided"
        self.project = "No information provided"
        self.DataInProject = "No information provided"
        self.DataPerFile = "No information provided"
        self.totalFYData = "No information provided"
        self.RunningDataTotal = "No information provided"

    def Print(self):
        print("""
        ID: {0}
        ObjType: {1}
        Name: {2}
        FY: {3}
        Project: {4}
        DataInProj: {5}
        DataPerFile: {6}
        TotalFYdata: {7}
        RunningTotal: {8}
        """.format(self.ID, self.ObjectType, self.name, self.FY, self.project, 
                   self.DataInProject, self.DataPerFile, self.totalFYData, 
                   self.RunningDataTotal)
        )

    def toJSON(self):
        return(json.dumps(self, default=lambda o: o.__dict__, 
            sort_keys=True, indent=4))



class problemSBitem(object):

    def __init__(self):
        self.ID = "No information provided"
        self.URL = "No information provided"
    
    def toJSON(self):
        return(json.dumps(self, default=lambda o: o.__dict__, 
            sort_keys=True, indent=4))

def format_to_array_of_arrays_For_Json(report_Dict, sbItemList):
    report_Array = []
    # heading_Array = ['ID', 'Object Type', 'Name', 'Fiscal Year', 'Project', 'Data in Project(GB)', 'Data per File (KB)', 'Fiscal Year Total Data (GB)', 'Running Data Total(GB)']
    # report_Array.append(heading_Array)
    # item_Array = []
    num = 0
    for i in sbItemList:
        # item_Array[:] = []
        report_Array.append([])
        print("Current item object:")
        i.Print()
        print("report_Array new added: {0}".format(report_Array))
        report_Array[num].append(i.ID)
        report_Array[num].append(i.ObjectType)
        report_Array[num].append(i.name)
        report_Array[num].append(i.FY)
        report_Array[num].append(i.project)
        report_Array[num].append(i.DataInProject)
        report_Array[num].append(i.DataPerFile)
        report_Array[num].append(i.totalFYData)
        report_Array[num].append(i.RunningDataTotal)
        print("report_Array all added: {0}".format(report_Array))

        # report_Array.append(item_Array)
        print("report_Array after append: {0}".format(report_Array))
        num += 1
    print('Here\'s the new report_Array: ')
    print(report_Array)
    return(report_Array)


def formatForJson(report_Dict):
    sbItemList = []
    numItems = len(report_Dict["ID"])
    print("numItems")
    print(numItems)
    for i in range(0, numItems):
        x = sbItem()
        x.ID = report_Dict['ID'][i]
        x.ObjectType = report_Dict['Object Type'][i]
        x.name = report_Dict['Name'][i]
        x.FY = report_Dict['Fiscal Year'][i]
        x.project = report_Dict['Project'][i]
        x.DataInProject = report_Dict['Data in Project (GB)'][i]
        x.DataPerFile = report_Dict['Data per File (KB)'][i]
        x.totalFYData = report_Dict['Fiscal Year Total Data (GB)'][i]
        x.RunningDataTotal = report_Dict['Running Data Total (GB)'][i]
        sbItemList.append(x)

    print("sbItemList: ")
    print(sbItemList)
    for i in sbItemList:
        print("i.Print():")
        i.Print()
    # report_Array = format_to_array_of_arrays_For_Json(report_Dict, sbItemList)
    return(sbItemList)
    # return(report_Array)
    
def format_Missing_and_Exceptions(IDlist, URLlist):
    outputList = []
    numItems = len(IDlist)
    for i in range(0, numItems):
        x = problemSBitem()
        x.ID = IDlist[i]
        x.URL = URLlist[i]
        outputList.append(x)
    return(outputList)


    # for key, value in report_Dict.items():
    #     print("Key:")  # Quantico
    #     print(key)  # Quantico
    #     print("Value:")  # Quantico
    #     print(value)  # Quantico
    #     for i in report_Dict[key]:
    #         print(report_Dict[key])  # Quantico
    #         print(report_Dict[key][place])  # Quantico
    #         print("Length:")  # Quantico
    #         print(len(report_Dict[key]))  # Quantico


def main():
    #try:
    PossiblePermissionsIssuesURL = []
    PossiblePermissionsIssuesURL[:] = []
    MissingDataURL= []
    MissingDataURL[:] = []
    if gl.Exceptions != []:
        for i in gl.Exceptions:
            json = sb.get_item(i)
            PossiblePermissionsIssuesURL.append(json['link']['url'])
    if gl.MissingData != []:
        for i in gl.MissingData:
            json = sb.get_item(i)
            MissingDataURL.append(json['link']['url'])

    # wb = Workbook()
    # ws = wb.active
    # ws.title = "Report"

    # if gl.MissingData != []:
    #     ws_missing = wb.create_sheet("Missing", 1) #Inserts a new sheet, named "Missing" in the second position
    #     #ws_missing.sheet_properties.tabColor = "1072BA" #Makes the tab for the sheet red so it draws attention.
    # if gl.Exceptions != []:
    #     ws_exceptions = wb.create_sheet("Exceptions", 2) #Inserts a new sheet, named "Exceptions" in the third position
        #ws_exceptions.sheet_properties.tabColor = "1072BA" #Makes the tab for the sheet red so it draws attention.


    report_Dict = {'ID': gl.ID, 'Object Type': gl.ObjectType, 'Name': gl.Name,
                    'Fiscal Year': gl.FiscalYear, 'Project': gl.Project,
                    'Data in Project (GB)': gl.DataInProject,
                    'Data per File (KB)': gl.DataPerFile,
                    'Fiscal Year Total Data (GB)': gl.totalFYDataList,
                    'Running Data Total (GB)': gl.RunningDataTotal,
                        }
    print("report_Dict:")  # Quantico
    print(report_Dict)  # Quantico
                        #include these eventually: 'Missing Data?': L7MissingData, 'Exceptions/Permissions Issues': L9Exceptions

                #include these eventually: 'Missing Data?', 'Exceptions/Permissions Issues'
    # #if gl.MissingData != []:
    # report_Missing = {'Missing Data ID': gl.MissingData,
    #                         'Missing Data URL': MissingDataURL}

    # #if gl.Exceptions != []:
    # report_Exceptions = {'Exception/Permission Issue ID': gl.Exceptions,
    #                            'Exception/Permission Issue URL':
    #                            PossiblePermissionsIssuesURL} 
    #                            #include 'Exception ID': L10Exceptions_IDs, later
 #include 'Exception ID' later


    # for r in dataframe_to_rows(dfOrdered, index=False, header=True):
    #     ws.append(r)

    # for cell in ws[1]:
    #     cell.style = 'Pandas'

    print("before report")  # Quantico
    report = formatForJson(report_Dict)
    print(report)  # Quantico

    reportDict = {}
    reportDict['report'] = report

    WriteMissing = None
    WriteExceptions = None
    if MissingDataURL != []:
        WriteMissing = True
        report_Missing = format_Missing_and_Exceptions(gl.MissingData, 
                                                       MissingDataURL)
        reportDict['missing'] = report_Missing
    elif MissingDataURL == []:
        WriteMissing = False
        reportDict['missing'] = 'None'
    else:
        print('Something went wrong when transfering report_Missing to reportDict in ExcelPring.py')
        # for r in dataframe_to_rows(df_missing, index=False, header=True):
        #     ws_missing.append(r)

        # for cell in ws_missing[1]:
        #     cell.style = 'Pandas'

    if gl.Exceptions != []:
        WriteExceptions = True
        report_Exceptions = format_Missing_and_Exceptions(gl.Exceptions,
                                                PossiblePermissionsIssuesURL)
        reportDict['exceptions'] = report_Exceptions
    elif gl.Exceptions == []:
        WriteExceptions = False
        reportDict['exceptions'] = 'None'
    else:
        print('Something went wrong when transfering report_Exceptions to reportDict in ExcelPring.py')
        # for r in dataframe_to_rows(df_exceptions, index=False, header=True):
        #     ws_exceptions.append(r)

        # for cell in ws_exceptions[1]:
        #     cell.style = 'Pandas'

    #print(df)
    #print('''

#    ---------Let\'s try reordering that...

 #   ''')
    #print(dfOrdered)
    #flash(dfOrdered['ID'])
    #flash(dfOrdered['Object Type'])
    #flash(dfOrdered['Name'])
    #flash(dfOrdered['Fiscal Year'])
    #flash(dfOrdered['Project'])
    #flash(dfOrdered['Data in Project (GB)'])
    #flash(dfOrdered['Data per File (KB)'])
    #flash(dfOrdered['Fiscal Year Total Data (GB)'])
    #flash(dfOrdered['Running Data Total (GB)'])
    #if gl.Exceptions != [] or gl.MissingData != []:
    #    print('''

    #Items missing data:
    #          ''')
    #    print(df_missing)
    #    print('''

    #Exceptions raised:
    #          ''')
    #    print(df_exceptions)
    #print('''
    #Does that look correct?
    #(Y / N)
    #''')
    #correct = input("> ").lower()
    #if 'y' in correct:
    #    ask(wb)
    #elif 'n' in correct:
    #    import editGPY
    #    editGPY.main()
    #    main()
    #ask(wb)
    #download_log(wb, dfOrdered, df_missing, df_exceptions, WriteMissing, WriteExceptions)
    #saveExcel(wb)
    # reportDict = createJson(wb, dfOrdered, df_missing, df_exceptions, WriteMissing, WriteExceptions)
    pprint(reportDict)  # Quantico
    return(reportDict)

# def createJson(wb, dfOrdered, df_missing, df_exceptions, WriteMissing, WriteExceptions):
#     reportDict = {}
#     #print(dfOrdered)
#     report = dfOrdered.to_json()
#     #pprint(report)
#     reportDict['report'] = report
#     #pprint(reportDict)
#     if WriteMissing:
#         missing = df_missing.to_json()
#         reportDict['missing'] = missing
#     if WriteExceptions:
#         exceptions = df_exceptions.to_json()
#         reportDict['exceptions'] = exceptions
#     return(reportDict)

# def download_log(wb, dfOrdered, df_missing, df_exceptions, WriteMissing, WriteExceptions):
#    output = io.BytesIO()
#    writer = pandas.ExcelWriter(output, engine='xlsxwriter')
#    dfOrdered.to_excel(writer, sheet_name='Report')
#    if WriteMissing:
#         df_missing.to_excel(writer, sheet_name='Missing')
#    if WriteExceptions:
#        df_exceptions.to_excel(writer, sheet_name='Exceptions')
#    writer.save()
#    output.seek(0)
#    excelDownload = output.read()
#    #print(excelDownload)
#    return send_file(excelDownload, mimetype='application/vnd.ms-excel.sheet.binary.macroenabled.12', attachment_filename = 'sbMACRO-output.xlsx',
#                     as_attachment=True)

# def new_download_log(wb, dfOrdered, df_missing, df_exceptions, WriteMissing, WriteExceptions):
#    output = io.BytesIO()
#    writer = pandas.ExcelWriter(output, engine='xlsxwriter')
#    dfOrdered.to_excel(writer, sheet_name='Report')
#    if WriteMissing:
#         df_missing.to_excel(writer, sheet_name='Missing')
#    if WriteExceptions:
#        df_exceptions.to_excel(writer, sheet_name='Exceptions')
#    writer.save()
#    output.seek(0)
#    excelDownload = output.read()
#    return send_file(excelDownload, mimetype='application/vnd.ms-excel.sheet.binary.macroenabled.12', attachment_filename = 'sbMACRO-output.xlsx',
#                     as_attachment=True)

# def ask(wb):
#     print ('''

#     Would you like to save this?
#     (Y / N)''')
#     answer = input('> ').lower()
#     if 'y' in answer:
#         print('''
#     Where would you like to save the file? Copy and paste a file path or '''+
#     '''type "Desktop" to save it to the desktop (If you run Linux, you must'''+
#     ''' paste a path).''')
#         answer2 = input("> ").lower()
#         if 'desk' in answer2:
#             filePath = os.path.expanduser("~/Desktop/")
#             saveExcel(wb, filePath)
#         elif '/' in answer2 or '\\' in answer2:
#             filePath = answer2
#             saveExcel(wb, filePath)
#         else:
#             print('''
#     That didn't appear to be a path. Please try again.''')
#             ask(wb)

#     elif 'n' in answer:
#         print('''
#         Ok, we won't save it.''')
#         return
# #    except (ValueError, Exception) as e:
# #        print('''
# #    ----------------------------WARNING: Something went wrong in the function "main" in ExcelPrint.py.''')
# #        exit()

# def saveExcel(wb):
#     filepath = "C:/Users/Taylor/Documents/!USGS/Python/sbProgramGitRepo/TrialWebApp/static/"   # Quantico, eyekeeper, must change this in final hosted version.
#     name = 'sbMACRO-output.xlsx'
#     wb.save(filepath+name)
#     print("file saved")
#     return send_from_directory(filepath, name, mimetype='application/vnd.ms-excel.sheet.binary.macroenabled.12', as_attachment=True)

# def saveExcel_old(wb, filePath):
#     ChosenFiscalYear = gl.FiscalYear[-1]  # Most recent Fiscal Year.
#     print('''
#     Would you like to name it something other than "'''+str(ChosenFiscalYear)+
#     ''' Data Metrics.xlsx"?
#     Beware: chosing the same name as an existing file overwrites that file '''+
#     '''without warning.''')
#     answer = input('> ').lower()
#     if 'y' in answer or 'other' in answer:
#         print('''
#     What name would you like to give the Excel file?''')
#         name = input('> ')
#         wb.save(str(filePath)+str(name)+".xlsx")
#         print('''
#     Workbook saved as "'''+str(name)+'''.xlsx" in \''''+str(filePath)+'''\'.''')
#         return
#     elif 'n' in answer or 'keep' in answer:
#         wb.save(str(filePath)+str(ChosenFiscalYear)+" Data Metrics.xlsx")
#         print('''
#     Workbook saved as "'''+str(ChosenFiscalYear)+''' Data Metrics.xlsx" in \''''+str(filePath)+'''\'.''')
#         return
#     else:
#         print('''
#     I'm sorry, I didn't get that. Please type 'y', 'other', 'n', or 'keep'
#     ''')
#         saveExcel(wb, filePath)


if __name__ == '__main__':
    main()
